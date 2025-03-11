import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, time, timedelta
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment
import sys
import os
import json
import numpy as np
import threading

# 配置文件相关路径
script_path = os.path.abspath(sys.argv[0])
script_dir = os.path.dirname(script_path)
script_name = os.path.splitext(os.path.basename(script_path))[0]
CONFIG_FILE = os.path.join(script_dir, f"{script_name}_config.json")


def generate_time_options():
    """生成时间下拉选项（每30分钟间隔）"""
    return [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 30)]


def excel_column_to_number(col_letter):
    """将Excel列字母转换为数字"""
    try:
        return column_index_from_string(col_letter) - 1
    except ValueError as e:
        raise ValueError(f"无效的列标识: {col_letter}") from e


def number_to_excel_column(n):
    """将数字转换为Excel列字母"""
    return get_column_letter(n + 1)


def format_time(time_value, time_format):
    """根据时间格式返回对应的时间字符串"""
    if time_value is None or pd.isnull(time_value):
        return ""

    if time_format == "小时时间格式":
        return round(time_value, 2)
    elif time_format == "复合时间格式":
        total_minutes = int(time_value * 60)
        days = total_minutes // 1440
        remaining_minutes = total_minutes % 1440
        hours = remaining_minutes // 60
        minutes = remaining_minutes % 60

        time_parts = []
        if days > 0:
            time_parts.extend([f"{days}天", f"{hours}小时", f"{minutes}分钟"])
        elif hours > 0:
            time_parts.extend([f"{hours}小时", f"{minutes}分钟"])
        elif minutes > 0:
            time_parts.append(f"{minutes}分钟")
        return " ".join(time_parts) if time_parts else "0"
    else:
        return time_value


class ConfigWindow:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("工时计算配置 v5.0")
        self.root.minsize(480, 600)
        self.final_config = None
        self.processing_done = False
        self.original_file_path = ""
        self.original_sheet_name = ""
        self.time_options = generate_time_options()
        self.time_slots = []
        self.open_dir_var = tk.BooleanVar(value=True)  # 新增：打开目录复选框变量
        self.topmost_var = tk.BooleanVar(value=True)  # 新增：置顶窗口复选框变量

        # GUI样式配置
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self._configure_styles()
        self._create_widgets()
        self.load_config_from_file()
        self.root.protocol("WM_DELETE_WINDOW", self.safe_exit)

        # 设置窗口置顶状态
        self.update_topmost_state()

    def _configure_styles(self):
        """样式配置"""
        self.style.configure(".", background="#F5F5F5", font=("微软雅黑", 9))
        self.style.configure("TLabel", anchor="e")
        self.style.configure("Red.TLabel", foreground="red")
        self.style.configure("Green.TLabel", foreground="#4CAF50")
        self.style.configure(
            "Accent.TButton",
            foreground="white",
            background="#2196F3",
            borderwidth=0,
            font=("微软雅黑", 10, "bold"),
        )
        self.style.map(
            "Accent.TButton",
            background=[("active", "#1976D2"), ("disabled", "#BBDEFB")],
        )
        self.style.configure("Red.TLabelframe", bordercolor="red")

    def _create_widgets(self):
        """创建主界面组件"""
        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.pack(expand=True, fill=tk.BOTH)

        self._create_file_section(main_frame)
        self._create_sheet_section(main_frame)
        self._create_column_section(main_frame)
        self._create_time_slot_section(main_frame)
        self._create_time_format_section(main_frame)
        self._create_checkbox_section(main_frame)
        self._create_status_bar(main_frame)
        self._create_action_buttons(main_frame)

    def _create_file_section(self, parent):
        """文件选择区域"""
        frame = ttk.LabelFrame(parent, text=" 文件配置 ", padding=10)
        frame.pack(fill=tk.X, pady=5)

        self.file_entry = ttk.Entry(frame)
        self.file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        self.browse_btn = ttk.Button(frame, text="浏览...", command=self.select_file)
        self.browse_btn.pack(side=tk.RIGHT, padx=5)

    def _create_sheet_section(self, parent):
        """工作表选择区域"""
        frame = ttk.LabelFrame(parent, text=" 表名设置 ", padding=10)
        frame.pack(fill=tk.X, pady=5)

        self.sheet_combobox = ttk.Combobox(frame, state="readonly")
        self.sheet_combobox.pack(fill=tk.X, padx=5, pady=3)

    def _create_column_section(self, parent):
        """列配置区域"""
        frame = ttk.LabelFrame(parent, text=" 列配置 ", padding=10)
        frame.pack(fill=tk.X, pady=5)

        config_fields = [
            ("* 开始时间列标:", "start_col", "请输入字母，如A或AA"),
            ("* 结束时间列标:", "end_col", "请输入字母，如B或AB"),
            ("  写入时长列标:", "write_col", "如留空则为结束时间列右侧"),
            ("* 计算起始行号:", "start_row", "执行计算的起始行"),
        ]

        self.entries = {}
        for label, key, help_text in config_fields:
            row_frame = ttk.Frame(frame)
            row_frame.pack(fill=tk.X, pady=2)

            ttk.Label(
                row_frame,
                text=label,
                width=12,
                style="Red.TLabel" if "*" in label else "TLabel",
            ).pack(side=tk.LEFT)

            entry = ttk.Entry(row_frame, width=12)
            entry.pack(side=tk.LEFT, padx=2)
            self.entries[key] = entry

            ttk.Label(row_frame, text=help_text, foreground="#666", width=36).pack(
                side=tk.LEFT, padx=5
            )

        self.entries["start_col"].insert(0, "A")
        self.entries["end_col"].insert(0, "B")
        self.entries["start_row"].insert(0, "2")

    def _create_time_slot_section(self, parent):
        """工作时间段设置区域"""
        frame = ttk.LabelFrame(parent, text=" 工作时间段设置 ", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.time_slots_container = ttk.Frame(frame)
        self.time_slots_container.pack(fill=tk.BOTH, expand=True)

        if not hasattr(self, "time_slots_loaded"):
            self.add_time_slot(("08:30", "12:00"))
            self.add_time_slot(("13:30", "18:00"))

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=5)
        ttk.Button(btn_frame, text="+ 添加时间段", command=self.add_time_slot).pack(
            side=tk.LEFT, padx=5
        )
        ttk.Button(
            btn_frame, text="- 删除最后一段", command=self.remove_last_slot
        ).pack(side=tk.LEFT)

    def _create_time_format_section(self, parent):
        """时间格式选择区域"""
        frame = ttk.LabelFrame(parent, text=" 时间格式 ", padding=10)
        frame.pack(fill=tk.X, pady=5)

        ttk.Label(frame, text="显示格式:", width=14, anchor="e").pack(
            side=tk.LEFT, padx=5
        )
        self.time_format_var = tk.StringVar()
        self.time_format_combobox = ttk.Combobox(
            frame,
            textvariable=self.time_format_var,
            values=("小时时间格式", "复合时间格式"),
            state="readonly",
            width=18,
        )
        self.time_format_combobox.pack(side=tk.LEFT)
        self.time_format_combobox.set("小时时间格式")

    def _create_checkbox_section(self, parent):
        """自动保存设置"""
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=5)

        self.auto_save_var = tk.BooleanVar()
        self.auto_save_check = ttk.Checkbutton(
            frame, text="自动保存配置", variable=self.auto_save_var
        )
        self.auto_save_check.pack(side=tk.LEFT)

        self.day_calc_var = tk.BooleanVar()
        self.day_calc_check = ttk.Checkbutton(
            frame,
            text="按一天24小时计算",
            variable=self.day_calc_var,
            command=self.toggle_day_calc,
        )  # 绑定勾选事件
        self.day_calc_check.pack(side=tk.LEFT, padx=10)

        # 新增：打开目录复选框
        self.open_dir_check = ttk.Checkbutton(
            frame, text="处理完成后打开文件目录", variable=self.open_dir_var
        )
        self.open_dir_check.pack(side=tk.LEFT, padx=10)

        # 新增：置顶窗口复选框
        self.topmost_check = ttk.Checkbutton(
            frame,
            text="窗口始终置顶",
            variable=self.topmost_var,
            command=self.toggle_topmost,
        )
        self.topmost_check.pack(side=tk.LEFT, padx=10)

    def _create_status_bar(self, parent):
        """状态栏"""
        self.status_label = ttk.Label(parent, text="就绪", foreground="#666")
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X, pady=5)

    def _create_action_buttons(self, parent):
        """操作按钮区域"""
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=tk.X, pady=10)

        self.load_btn = ttk.Button(
            btn_frame, text="加载配置", command=self.load_config_dialog
        )
        self.load_btn.pack(side=tk.LEFT, padx=5)
        self.save_btn = ttk.Button(
            btn_frame, text="保存配置", command=self.save_config_dialog
        )
        self.save_btn.pack(side=tk.LEFT, padx=5)
        self.ok_btn = ttk.Button(
            btn_frame,
            style="Accent.TButton",
            text="开始计算",
            command=self.validate_inputs,
        )
        self.ok_btn.pack(side=tk.RIGHT, padx=5)

    def add_time_slot(self, default=("", "")):
        """添加工作时间段输入行（带独立错误提示）"""
        frame = ttk.Frame(self.time_slots_container)
        frame.pack(fill=tk.X, pady=2, padx=5)

        ttk.Label(frame, text="*", style="Red.TLabel").pack(side=tk.LEFT)

        time_frame = ttk.Frame(frame)
        time_frame.pack(side=tk.LEFT, padx=3)

        ttk.Label(time_frame, text="开始：").pack(side=tk.LEFT)

        start_var = tk.StringVar(value=default[0])
        start_cb = ttk.Combobox(
            time_frame,
            textvariable=start_var,
            values=self.time_options,
            width=7,
            state="readonly",
        )
        start_cb.pack(side=tk.LEFT)

        ttk.Label(time_frame, text="  结束：").pack(side=tk.LEFT)

        end_var = tk.StringVar(value=default[1])
        end_cb = ttk.Combobox(
            time_frame,
            textvariable=end_var,
            values=self.time_options,
            width=7,
            state="readonly",
        )
        end_cb.pack(side=tk.LEFT)

        error_label = ttk.Label(
            frame, foreground="red", font=("微软雅黑", 9), wraplength=300
        )
        error_label.pack(side=tk.RIGHT, padx=5)

        for cb in [start_cb, end_cb]:
            cb.bind("<<ComboboxSelected>>", lambda e: self.validate_time_slots())

        self.time_slots.append(
            {"start": start_var, "end": end_var, "error": error_label}
        )
        self.validate_time_slots()

    def validate_time_slots(self):
        """验证时间段设置（带独立错误提示）"""
        valid_periods = []
        all_errors = []
        overlap_errors = []

        for idx, slot in enumerate(self.time_slots):
            start = slot["start"].get().strip()
            end = slot["end"].get().strip()
            error_label = slot["error"]
            errors = []

            if not start:
                errors.append("开始时间不能为空")
            if not end:
                errors.append("结束时间不能为空")
            if errors:
                error_label.config(text=" | ".join(errors), foreground="red")
                continue

            try:
                start_time = datetime.strptime(start, "%H:%M").time()
                end_time = datetime.strptime(end, "%H:%M").time()

                if start_time >= end_time:
                    errors.append("开始时间不能晚于结束时间")
                else:
                    valid_periods.append((start_time, end_time))
            except ValueError:
                errors.append("时间格式应为HH:MM")

            if errors:
                error_label.config(text=" | ".join(errors), foreground="red")
            else:
                error_label.config(text="✓ 格式有效", foreground="#4CAF50")

        if valid_periods:
            sorted_periods = sorted(valid_periods, key=lambda x: x[0])
            for i in range(1, len(sorted_periods)):
                prev_end = sorted_periods[i - 1][1]
                curr_start = sorted_periods[i][0]
                if curr_start < prev_end:
                    overlap_errors.append(f"时间段 {i} 与 {i+1} 存在重叠")

            for idx, slot in enumerate(self.time_slots):
                if overlap_errors and idx < len(overlap_errors):
                    slot["error"].config(text=overlap_errors[idx], foreground="red")
                elif not slot["error"].cget("text").startswith("✓"):
                    continue

        return len(all_errors) == 0 and len(overlap_errors) == 0

    def remove_last_slot(self):
        """删除最后一个时间段"""
        if len(self.time_slots) > 1:
            last_slot = self.time_slots.pop()
            last_slot["error"].master.destroy()
            self.validate_time_slots()

    def toggle_controls(self, state):
        """启用或禁用控件"""
        widgets = [
            self.file_entry,
            self.browse_btn,
            self.sheet_combobox,
            self.load_btn,
            self.save_btn,
            self.ok_btn,
            *self.entries.values(),
            self.auto_save_check,
            self.time_format_combobox,
        ]
        for widget in widgets:
            widget.configure(state=state)
        self.status_label.configure(
            foreground="#666" if state == tk.NORMAL else "#2196F3"
        )

    def safe_exit(self):
        """安全退出程序"""
        self.root.destroy()

    def load_config_from_file(self):
        """从配置文件加载配置（新增状态同步）"""
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r") as f:
                    config = json.load(f)

                self.original_file_path = config.get("file_path", "")
                self.original_sheet_name = config.get("sheet_name", "")

                self.file_entry.delete(0, tk.END)
                self.file_entry.insert(0, self.original_file_path)

                if self.original_file_path:
                    self.load_sheets(self.original_file_path)

                for child in self.time_slots_container.winfo_children():
                    child.destroy()
                self.time_slots.clear()

                work_periods = config.get(
                    "work_periods", [["08:30", "12:00"], ["13:30", "18:00"]]
                )
                for period in work_periods:
                    self.add_time_slot(period)

                for key in self.entries:
                    self.entries[key].delete(0, tk.END)
                    self.entries[key].insert(0, config.get(key, ""))
                self.auto_save_var.set(config.get("auto_save", False))
                self.time_format_var.set(config.get("time_format", "小时时间格式"))
                self.day_calc_var.set(config.get("day_calc", False))
                self.open_dir_var.set(config.get("open_dir", True))  # 加载打开目录设置
                self.topmost_var.set(config.get("topmost", True))  # 加载置顶设置

                # 强制设置表名（新增修复点）
                if self.original_sheet_name:
                    self.sheet_combobox.set(self.original_sheet_name)

                # 新增：加载配置后同步UI状态
                self.toggle_day_calc()

            except Exception as e:
                messagebox.showerror("配置错误", f"加载失败: {str(e)}")

    def get_current_config(self):
        """获取当前配置"""
        return {
            "file_path": self.file_entry.get().strip(),
            "sheet_name": self.sheet_combobox.get().strip(),
            "start_col": self.entries["start_col"].get().strip().upper(),
            "end_col": self.entries["end_col"].get().strip().upper(),
            "write_col": self.entries["write_col"].get().strip().upper(),
            "start_row": self.entries["start_row"].get().strip(),
            "auto_save": self.auto_save_var.get(),
            "time_format": self.time_format_var.get(),
            "day_calc": self.day_calc_var.get(),
            "work_periods": [
                [slot["start"].get(), slot["end"].get()]
                for slot in self.time_slots
                if slot["start"].get() and slot["end"].get()
            ],
            "open_dir": self.open_dir_var.get(),  # 新增：保存打开目录的设置
            "topmost": self.topmost_var.get(),  # 新增：保存置顶设置
        }

    def save_config_to_file(self, config, silent=False):
        """将配置保存到文件"""
        try:
            with open(CONFIG_FILE, "w") as f:
                json.dump(config, f, indent=2)
            if not silent:
                messagebox.showinfo("保存成功", f"配置已保存到：\n{CONFIG_FILE}")
            return True
        except Exception as e:
            if not silent:
                messagebox.showerror("保存错误", str(e))
            return False

    def validate_inputs(self):
        """验证输入"""
        try:
            config = self.get_current_config()
            errors = []

            if not config["file_path"]:
                errors.append("请选择Excel文件")
            elif not os.path.exists(config["file_path"]):
                errors.append("文件路径不存在")

            for col in [config["start_col"], config["end_col"]]:

                if not col.isalpha():
                    errors.append("列标识必须为字母")

            try:
                start_row = int(config["start_row"])
                if start_row < 1:
                    errors.append("起始行号必须≥1")
            except:
                errors.append("起始行号格式错误")

            write_col = config["write_col"]
            if write_col:
                if not write_col.isalpha():
                    errors.append("写值列标识必须为字母")
                elif write_col == config["start_col"] or write_col == config["end_col"]:
                    errors.append("写值列不能与开始/结束列相同")

            if not self.validate_time_slots():
                errors.append("请修正时间段设置错误")

            if errors:
                raise ValueError("\n".join(errors))

            self.final_config = {
                "file_path": config["file_path"],
                "sheet_name": config["sheet_name"] or None,
                "start_col": excel_column_to_number(config["start_col"]),
                "end_col": excel_column_to_number(config["end_col"]),
                "write_col": excel_column_to_number(write_col) if write_col else None,
                "skiprows": int(config["start_row"]) - 1,
                "auto_save": self.auto_save_var.get(),
                "time_format": config["time_format"],
                "day_calc": config["day_calc"],
                "work_periods": [
                    (
                        datetime.strptime(p[0], "%H:%M").time(),
                        datetime.strptime(p[1], "%H:%M").time(),
                    )
                    for p in config["work_periods"]
                ],
                "open_dir": config["open_dir"],  # 新增：传递打开目录设置
            }

            self.toggle_controls(tk.DISABLED)
            self.processing_done = False
            self.status_label.configure(text="处理中...")

            processing_thread = threading.Thread(target=self.run_processing)
            processing_thread.start()
            self.monitor_processing(processing_thread)

        except Exception as e:
            messagebox.showerror("输入错误", str(e))
            self.toggle_controls(tk.NORMAL)

    def monitor_processing(self, thread):
        """监控处理线程"""
        if thread.is_alive():
            self.root.after(100, lambda: self.monitor_processing(thread))
        else:
            self.processing_done = True

    def run_processing(self):
        """运行处理"""
        try:
            success, result_data = main_process(self.final_config)
            if success:
                self.root.after(0, lambda: self.show_result(result_data))
                if self.final_config["auto_save"]:
                    current_config = self.get_current_config()
                    self.save_config_to_file(current_config, silent=True)
            else:
                self.root.after(0, self.handle_processing_failure)
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("处理错误", str(e)))
            self.root.after(0, self.handle_processing_failure)

    def handle_processing_failure(self):
        """处理失败时的恢复操作"""
        self.toggle_controls(tk.NORMAL)
        self.status_label.configure(text="处理失败，请检查配置")

    def show_result(self, result_data):
        """显示处理结果"""
        self.status_label.configure(text="处理完成")
        messagebox.showinfo("处理完成", "\n".join(result_data))
        self.toggle_controls(tk.NORMAL)

        # 根据复选框状态决定是否打开文件目录
        if self.final_config["open_dir"]:
            self.open_result_directory()

    def open_result_directory(self):
        """打开结果目录"""
        dir_path = os.path.dirname(self.final_config["file_path"])
        if os.path.exists(dir_path):
            os.startfile(dir_path)

    def select_file(self):
        """选择文件"""
        file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if file_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
            self.load_sheets(file_path)

    def load_sheets(self, file_path):
        """加载Excel文件的表名（修复表名不存在处理逻辑）"""
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            self.sheet_combobox["values"] = sheet_names

            if not sheet_names:
                self.sheet_combobox.set("")
                messagebox.showwarning("空工作表", "该Excel文件没有工作表")
                return

            if file_path == self.original_file_path:
                # 修复点：当原配置表名不存在时强制切换首表
                if (
                    self.original_sheet_name
                    and self.original_sheet_name not in sheet_names
                ):
                    new_sheet = sheet_names[0]
                    message = f"原配置表名 '{self.original_sheet_name}' 不存在\n已自动切换为 '{new_sheet}'"
                    messagebox.showwarning("表名变更", message)
                    self.sheet_combobox.set(new_sheet)
                    self.original_sheet_name = new_sheet  # 更新原始表名记录
                else:
                    self.sheet_combobox.set(self.original_sheet_name or sheet_names[0])
            else:
                self.sheet_combobox.set(sheet_names[0])
                self.original_sheet_name = sheet_names[0]  # 新文件时重置原始表名

        except Exception as e:
            messagebox.showerror("加载表名错误", str(e))

    def save_config_dialog(self):
        """保存配置对话框"""
        self.save_config_to_file(self.get_current_config())

    def load_config_dialog(self):
        """加载配置对话框（修复表名加载逻辑）"""
        file_path = filedialog.askopenfilename(filetypes=[("JSON配置", "*.json")])
        if file_path:
            try:
                with open(file_path, "r") as f:
                    config = json.load(f)

                self.original_file_path = config.get("file_path", "")
                self.original_sheet_name = config.get("sheet_name", "")

                # 清空现有时间段
                for child in self.time_slots_container.winfo_children():
                    child.destroy()
                self.time_slots.clear()

                work_periods = config.get(
                    "work_periods", [["08:30", "12:00"], ["13:30", "18:00"]]
                )
                for period in work_periods:
                    self.add_time_slot(period)

                for key in self.entries:
                    self.entries[key].delete(0, tk.END)
                    self.entries[key].insert(0, config.get(key, ""))
                self.auto_save_var.set(config.get("auto_save", False))
                self.time_format_var.set(config.get("time_format", "小时时间格式"))
                self.day_calc_var.set(config.get("day_calc", False))
                self.open_dir_var.set(
                    config.get("open_dir", True)
                )  # 新增：加载打开目录设置
                self.topmost_var.set(config.get("topmost", True))  # 新增：加载置顶设置

                # 强制设置表名（新增修复点）
                if self.original_sheet_name:
                    self.sheet_combobox.set(self.original_sheet_name)

                # 触发工作表加载
                if self.original_file_path:
                    self.file_entry.delete(0, tk.END)
                    self.file_entry.insert(0, self.original_file_path)
                    self.load_sheets(self.original_file_path)

            except Exception as e:
                messagebox.showerror("配置错误", f"加载失败: {str(e)}")

    def toggle_day_calc(self):
        """切换按一天24小时计算时的逻辑（完整修复版本）"""
        style = ttk.Style()
        if self.day_calc_var.get():

            # 修改标题样式
            self.time_slots_container.master.configure(
                text=" 工作时间段设置（按天计算模式已启用，时间段无效）",
                style="Red.TLabelframe",
            )
            style.configure("Red.TLabelframe.Label", foreground="red")
            self.time_slots_container.master.configure(style="Red.TLabelframe")
        else:
            # 恢复标题样式
            self.time_slots_container.master.configure(
                text=" 工作时间段设置 ", style="TLabelframe"
            )
            style.configure("TLabelframe.Label", foreground="black")
            self.time_slots_container.master.configure(style="TLabelframe")

            # 启用操作按钮
            btn_frame = self.time_slots_container.master.master.winfo_children()[-2]
            for btn in btn_frame.winfo_children():
                if isinstance(btn, ttk.Button):
                    btn.configure(state=tk.NORMAL)

    def toggle_topmost(self):
        """切换窗口置顶状态"""
        self.update_topmost_state()

    def update_topmost_state(self):
        """更新窗口置顶状态"""
        self.root.attributes("-topmost", self.topmost_var.get())


def calculate_working_hours_vectorized(
    starts, ends, time_format, work_periods, day_calc
):
    """计算工作小时数（动态时间段版本）"""
    total_hours = np.empty(len(starts), dtype=object)
    error_stats = {"空值记录": 0, "格式错误": 0, "时间倒置": 0, "零值记录": 0}
    sunday_notes = {}  # 存储周日信息

    for i in range(len(starts)):
        sundays = []
        try:
            if pd.isnull(starts[i]) or pd.isnull(ends[i]):
                error_stats["空值记录"] += 1
                total_hours[i] = np.nan
                continue

            start_time = pd.to_datetime(starts[i], errors="coerce")
            end_time = pd.to_datetime(ends[i], errors="coerce")

            if pd.isnull(start_time) or pd.isnull(end_time):
                error_stats["格式错误"] += 1
                total_hours[i] = np.nan
                continue

            if start_time >= end_time:
                error_stats["时间倒置"] += 1
                total_hours[i] = np.nan
                continue

            current = start_time.to_pydatetime()
            end_dt = end_time.to_pydatetime()
            valid_hours = 0.0

            temp_day = current.replace(hour=0, minute=0, second=0, microsecond=0)
            end_day = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)

            while temp_day <= end_day:
                if temp_day.weekday() == 6:
                    sundays.append(temp_day.strftime("%m-%d"))
                temp_day += timedelta(days=1)

            if day_calc:
                # 按一天24小时计算
                valid_hours = (end_dt - current).total_seconds() / 3600
                # 排除周日的时间
                temp = current
                while temp <= end_dt:
                    if temp.weekday() == 6:
                        sunday_start = temp.replace(
                            hour=0, minute=0, second=0, microsecond=0
                        )
                        sunday_end = temp.replace(
                            hour=23, minute=59, second=59, microsecond=999999
                        )
                        # 计算周日的时间差
                        sunday_duration = min(end_dt, sunday_end) - max(
                            current, sunday_start
                        )
                        valid_hours -= sunday_duration.total_seconds() / 3600
                    temp += timedelta(days=1)
            else:
                # 按时间段计算
                while current < end_dt:
                    if current.weekday() == 6:
                        current = current.replace(hour=8, minute=30) + timedelta(days=1)
                        continue

                    day_start = current.replace(
                        hour=0, minute=0, second=0, microsecond=0
                    )
                    day_end = day_start + timedelta(days=1)

                    for start_t, end_t in work_periods:
                        period_start = day_start.replace(
                            hour=start_t.hour, minute=start_t.minute
                        )
                        period_end = day_start.replace(
                            hour=end_t.hour, minute=end_t.minute
                        )

                        overlap_start = max(current, period_start)
                        overlap_end = min(end_dt, period_end)

                        if overlap_start < overlap_end:
                            delta = (overlap_end - overlap_start).total_seconds() / 3600
                            valid_hours += delta

                    current = day_end

            if valid_hours < 0:
                valid_hours = 0.0
                error_stats["零值记录"] += 1
            elif valid_hours == 0:
                error_stats["零值记录"] += 1

            total_hours[i] = valid_hours

            if sundays:
                sunday_notes[i] = sundays

        except Exception:
            error_stats["格式错误"] += 1
            total_hours[i] = np.nan

    formatted_hours = []
    for hours in total_hours:
        if pd.isnull(hours):
            formatted_hours.append(np.nan)
        else:
            formatted_hours.append(format_time(hours, time_format))

    return pd.Series(formatted_hours).astype(object), error_stats, sunday_notes


def main_process(config):
    """主处理函数"""
    try:
        sheet_name = config.get("sheet_name", None)
        display_sheet_name = sheet_name if sheet_name else "活动工作表"

        df = pd.read_excel(
            config["file_path"],
            sheet_name=sheet_name,
            header=None,
            skiprows=config["skiprows"],
            usecols=[config["start_col"], config["end_col"]],
            names=["start_time", "end_time"],
            engine="openpyxl",
        )

        wb = load_workbook(config["file_path"])
        if sheet_name is not None:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if config["write_col"] is not None:
            insert_col = config["write_col"]
        else:
            insert_col = config["end_col"] + 1

        target_col = get_column_letter(insert_col + 1)
        conflict_range = f"{target_col}{config['skiprows']+1}:{target_col}{ws.max_row}"
        conflict_values = [cell[0].value for cell in ws[conflict_range]]

        if any(conflict_values):
            conflict_cells = []
            for i, val in enumerate(conflict_values, start=config["skiprows"] + 1):
                if val is not None:
                    conflict_cells.append(f"{target_col}{i}")
                    if len(conflict_cells) >= 3:
                        break

            error_msg = [
                f"■ 工作表：{display_sheet_name}",
                f"目标列 {target_col} 存在数据冲突：",
                f"发现 {sum(1 for v in conflict_values if v is not None)} 个非空单元格",
                f"示例：{', '.join(conflict_cells)}...",
                "\n请清空目标列或手动插入新列！",
            ]
            messagebox.showerror("数据冲突", "\n".join(error_msg))
            return False, None

        df["work_hours"], error_stats, sunday_notes = (
            calculate_working_hours_vectorized(
                df["start_time"],
                df["end_time"],
                config["time_format"],
                config["work_periods"],
                config["day_calc"],
            )
        )

        for i in range(len(df)):
            row_num = config["skiprows"] + 1 + i
            if not pd.isnull(df["work_hours"][i]):
                cell = ws.cell(
                    row=row_num, column=insert_col + 1, value=df["work_hours"][i]
                )
                cell.alignment = Alignment(horizontal="right")
                if i in sunday_notes:
                    comment_text = f"包含{len(sunday_notes[i])}个周日：{', '.join(sunday_notes[i])}"
                    cell.comment = Comment(comment_text, "系统提示")
            elif pd.isnull(df["start_time"][i]) or pd.isnull(df["end_time"][i]):
                ws.cell(row=row_num, column=insert_col + 1, value="")

        total = len(df)
        valid = sum(~pd.isnull(df["work_hours"]))

        result_msg = [
            "■ 处理结果统计 ■",
            f"工作表名称：{display_sheet_name}",
            f"总记录数：{total} 条",
            f"✓ 有效记录：{valid} 条（含0值）",
            f"○ 零值记录：{error_stats['零值记录']} 条",
            f"✗ 无效记录：{total - valid} 条",
            "■ 异常分布 ■",
            f"空值记录：{error_stats['空值记录']} 条",
            f"格式错误：{error_stats['格式错误']} 条",
            f"时间倒置：{error_stats['时间倒置']} 条",
            f"\n文件已保存：{os.path.basename(config['file_path'])} ({config['time_format']})",
        ]

        wb.save(config["file_path"])
        return True, result_msg

    except Exception as e:
        error_msg = [
            f"■ 发生错误的工作表：{display_sheet_name}",
            f"错误类型：{str(e)}",
            "\n■ 排查建议：",
            "1. 确认Excel文件未被其他程序打开",
            "2. 检查时间列格式是否为标准时间格式",
            "3. 确保目标列（计算结果列）为空",
            "4. 验证工作表结构是否符合要求",
        ]
        messagebox.showerror("运行错误", "\n".join(error_msg))
        return False, None


if __name__ == "__main__":
    config_window = ConfigWindow()
    config_window.root.mainloop()
